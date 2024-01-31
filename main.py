import undetected_chromedriver as uc
from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import pandas as pd
from datetime import datetime
import tkinter as tk
import customtkinter as ctk
from openpyxl.styles import Alignment, PatternFill, Font, NamedStyle, Border, Side
from openpyxl import load_workbook
from copy import copy
import webbrowser
from PIL import Image

def scrape_data(codes, file_name, data_t):
    global status_label
    print(datetime.now().time())
    options = webdriver.ChromeOptions()
    options.headless = True
    driver = uc.Chrome(options=options)
    try:
        head = [["", "2018-2019", "2019-2020", "2020-2021", "2021-2022", "2022-2023", "2023-2024(Apr-May)", "2018-2019",
                 "2019-2020", "2020-2021", "2021-2022", "2022-2023", "2023-2024(Apr-May)", "2018-2019", "2019-2020",
                 "2020-2021", "2021-2022", "2022-2023", "2023-2024(Apr-May)"]]
        header = pd.DataFrame(head)
        data_t.lower()

        if data_t == "import":
            link = "https://tradestat.commerce.gov.in/eidb/icomcntq.asp"
        else:
            link = "https://tradestat.commerce.gov.in/eidb/ecomcntq.asp"

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            for code in codes:
                all_data = None  # Initialize as None
                i = 4
                while i >= 0:
                    driver.get(link)
                    # Wait for the year element to be present
                    try:
                        year_element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, '//select[@id="select2"]'))
                    )
                    except TimeoutException:
                        status_label.configure(text="Web server down, try again later")
                        return

                    year_select = Select(year_element)
                    year_select.select_by_index(i)  # Adjust index to start from 0

                    hs_code_input = driver.find_element(By.XPATH, '//input[@name="hscode"]')
                    hs_code_input.send_keys(code)

                    submit_btn = driver.find_element(By.XPATH, '//input[@value="Submit"]')
                    submit_btn.click()

                    # Wait for the table to be present
                    table = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//table[2]"))
                    )
                    cols = table.find_elements(By.XPATH, ".//th")
                    cols = [col.text.strip() for col in cols]

                    tableh = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//table[1]")))
                    colu = tableh.find_elements(By.XPATH, ".//th")
                    colu = [col.text.strip() for col in colu]
                    colls = []
                    for c in cols:
                        if '(' in c:
                            colls.append(c.split('(')[0])
                        else:
                            colls.append(c)
                    columns = colu[0:2] + [f"{col}(Values)" for col in colls[2:5]] + [f"{col}(Quantity)" for col in colls[5:]]

                    data_list = []
                    for row in table.find_elements(By.XPATH, './/tr'):
                        # Process each row and extract relevant information
                        cells = row.find_elements(By.XPATH, './/td')
                        # Your scraping logic here
                        data_list.append([cell.text.strip() for cell in cells])

                    # Create a DataFrame for the current table with specified column names
                    current_data = pd.DataFrame(data_list, columns=columns)

                    # Fill NaN values in the first DataFrame
                    current_data.fillna({col: 0 for col in current_data.columns}, inplace=True)

                    # Join the DataFrames based on the "Country / Region" column
                    if all_data is not None:
                        all_data = pd.merge(all_data, current_data, on='Country / Region', how='outer', suffixes=('_existing', f'_Year_{i}'))
                    else:
                        all_data = current_data
                    i -= 2

                # Drop the columns containing "%Growth" in their names and "S.No."
                columns_to_drop = ['%Growth', 'S.No.']
                all_data = all_data.loc[:, ~all_data.columns.str.contains('|'.join(columns_to_drop))]

                # Drop rows where "Country / Region" is numeric or "Total"
                all_data['Country / Region'] = all_data['Country / Region'].str.replace(',', '').str.strip()
                all_data = all_data[~pd.to_numeric(all_data['Country / Region'], errors='coerce').notna() | (all_data['Country / Region'] == 'Total')]
                all_data.drop(all_data[all_data['Country / Region']=='Total'].index, inplace=True)

                # Reorder columns: Values first, then Quantity
                value_cols = [col for col in all_data.columns if '(Values)' in col]
                quantity_cols = [col for col in all_data.columns if '(Quantity)' in col]
                value_cols = sorted(value_cols)
                quantity_cols = sorted(quantity_cols)
                all_data = all_data[['Country / Region'] + value_cols + quantity_cols]
                all_data = all_data.drop(all_data.index[0])
                all_data = all_data.reset_index(drop=True)
                all_data.replace("", 0, regex=True, inplace=True)
                all_data[value_cols] = all_data[value_cols].replace(',', '', regex=True).astype(float)
                all_data[quantity_cols] = all_data[quantity_cols].replace(',', '', regex=True).astype(float)

                # Save the joined DataFrame to a single Excel file
                years_col_names = set(col.split('(')[0][-9:] for col in all_data.columns if 'Values' in col)
                total_row = all_data.iloc[:, 1:].sum(axis=0)
                # Append the total row to the DataFrame
                total_df = pd.DataFrame({'Country / Region': ['Total'], **total_row.to_dict()}).fillna('')

                # Concatenate the original DataFrame and the total DataFrame
                all_data = pd.concat([all_data, total_df], ignore_index=True)
                for year_name in sorted(years_col_names):
                    value_col = f'{year_name}(Values)'
                    quantity_col = f'{year_name}(Quantity)'
                    price_col = f'{year_name}(Price)'
                    all_data[price_col] = all_data[value_col] * 10 ** 5 / all_data[quantity_col]
                all_data = all_data.round()
                all_data.to_excel(writer, sheet_name=str(code), index=False, startrow=2, header=False)
                print("Excel file updated")
                status_label.configure(text="Excel file updated")

        wb = load_workbook(file_name)
        for code in codes:
            ws = wb[str(code)]

            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
            ws.cell(row=1, column=2, value="Values in INR Lacs").alignment = Alignment(horizontal='center',
                                                                                       vertical='center')
            ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13)
            ws.cell(row=1, column=8, value="Quantity in MT").alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
            ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=19)
            ws.cell(row=1, column=14, value="Rates in INR/MT").alignment = Alignment(horizontal='center',
                                                                                     vertical='center')


            for col_num in range(1, 21):
                ws.column_dimensions[chr(ord('A') + col_num - 1)].width = 15
            # Set row height for the year rows (rows 2 and 3)
            row_heights = [20, 20]
            for row_num, height in enumerate(row_heights, 2):
                ws.row_dimensions[row_num].height = height
                ws.row_dimensions[row_num].width = 100
            # Write DataFrame to Excel
            border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
            for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=19):
                for cell in row:
                    cell.fill = PatternFill(start_color="004f71", end_color="004f71",
                                            fill_type="solid")  # Set the background color to yellow
                    cell.font = Font(color='ffffff', name="Gill Sans MT", bold=True, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border


            for r_idx, row in enumerate(header.values, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx + 1, column=c_idx, value=value)
            comma_style = NamedStyle(name=str(code), number_format='#,##0')
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=19):
                for cell in row:
                    cell.style = comma_style
                    cell.font = Font(name="Gill Sans MT", size=10)
                    cell.border = border
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            ws.cell(row=1, column=1, value="Country").alignment = Alignment(horizontal='center', vertical='center')

            source_range = ws['H1:S2']
            maxRow = ws.max_row
            dest_start = f'B{maxRow + 3}'
            total_start = f'B{maxRow + 5}'
            source_row = ws[f'H{ws.max_row}:S{ws.max_row}']
            for i, row in enumerate(source_row):
                for j, cell in enumerate(row):
                    # Create a new cell at the corresponding position in the destination range
                    dest_cell = ws[f"{chr(ord(total_start[0]) + j)}{int(total_start[1:]) + i}"]
                    dest_cell.value = cell.value
                    dest_cell.font = copy(cell.font)
                    dest_cell.number_format = copy(cell.number_format)

            # Iterate over the cells in the source range
            for i, row in enumerate(source_range):
                for j, cell in enumerate(row):
                    # Create a new cell at the corresponding position in the destination range
                    dest_cell = ws[f"{chr(ord(dest_start[0]) + j)}{int(dest_start[1:]) + i}"]
                    dest_cell.value = cell.value
                    if cell.has_style:
                        dest_cell.font = copy(cell.font)
                        dest_cell.border = copy(cell.border)
                        dest_cell.fill = copy(cell.fill)
                        dest_cell.alignment = copy(cell.alignment)
            # Save the workbook
            ws.merge_cells(start_row=maxRow + 3, start_column=2, end_row=maxRow + 3, end_column=7)
            ws.merge_cells(start_row=maxRow + 3, start_column=8, end_row=maxRow + 3, end_column=13)
            ws.merge_cells(start_row=maxRow + 3, start_column=14, end_row=maxRow + 3, end_column=19)
            # Save the Excel file
            wb.save(file_name)
            print("Excel file created successfully.")
            status_label.configure(text="Excel file created successfully")
    finally:
        print(datetime.now().time())
        driver.close()
        driver.quit()

def run_scraping():
    expiry = datetime(2024, 2, 26)
    if datetime.now() < expiry:
        codes = [int(code) for code in code_entry.get().split()]  # Convert space-separated codes to a list of integers
        file_name = file_name_entry.get()
        data_t = data_type_var.get()
        status_label.configure(text="Data Extraction in process, please wait")
        scrape_data(codes, file_name + ".xlsx", data_t)
    else:
        status_label.configure(text="Plan Expired")

def open_claude():
    webbrowser.open("https://claude.ai/chats")

# GUI setup

app = ctk.CTk(fg_color="gray10")
app.title("JVIA")
app.iconbitmap("Source/Icon1.ico")

logo_img = Image.open("Source/Icon1.png")
logo_photo = ctk.CTkImage(logo_img, size=(80,80))

# Create a frame for the tab buttons
tab_buttons_frame = ctk.CTkFrame(app)
tab_buttons_frame.place(x=1, y=1)

# Create a button for the first tab
data_extractor_button = ctk.CTkButton(tab_buttons_frame, text="Data Extractor", command=lambda: show_frame(data_extractor_frame, data_extractor_button)
                                      , corner_radius=0, width=225)
data_extractor_button.pack(side='left')

# Create a button for the second tab
gpt_tab_button = ctk.CTkButton(tab_buttons_frame, text="Claude AI", command=lambda: show_frame(claude_frame, gpt_tab_button),
                               corner_radius=0, width=225)
gpt_tab_button.pack(side='left')

logo = ctk.CTkLabel(app, image=logo_photo, text='')
logo.place(x=180, y=50)

# Function to show a frame
def show_frame(frame, button):
    data_extractor_frame.place_forget()
    claude_frame.place_forget()
    data_extractor_button.configure(fg_color="#b81d1d")
    gpt_tab_button.configure(fg_color="#b81d1d")
    frame.place(x=40, y=140)
    button.configure(fg_color="dark red")

# Create a frame for the existing elements
data_extractor_frame = ctk.CTkFrame(app, width=300, height=400, fg_color='gray10')

# Add your existing elements to data_extractor_frame here...
title_label = ctk.CTkLabel(data_extractor_frame, text="Export/Import Data Extractor", font=("", 25), text_color="#ffffff")
title_label.grid(row=1, column=0, columnspan=2, sticky='ew', pady=20)

# Code entry
code_label = ctk.CTkLabel(data_extractor_frame, text="Enter Codes (space-separated):", text_color="#ffffff")
code_label.grid(row=2, column=0, padx=10, pady=10)
code_entry = ctk.CTkEntry(data_extractor_frame)
code_entry.grid(row=2, column=1, padx=10, pady=10)

# File name entry
file_name_label = ctk.CTkLabel(data_extractor_frame, text="Enter File Name:", text_color="#ffffff")
file_name_label.grid(row=3, column=0, padx=10, pady=10)
file_name_entry = ctk.CTkEntry(data_extractor_frame)
file_name_entry.grid(row=3, column=1, padx=10, pady=10)

# Data type selection
data_type_label = ctk.CTkLabel(data_extractor_frame, text="Select Data Type:", text_color="#ffffff")
data_type_label.grid(row=4, column=0, padx=10, pady=10)
data_type_var = tk.StringVar(value="export")
data_type_combobox = ctk.CTkComboBox(data_extractor_frame, variable=data_type_var, values=["import", "export"])
data_type_combobox.grid(row=4, column=1, padx=10, pady=10)

# Run button
run_button = ctk.CTkButton(data_extractor_frame, text="Get Data", command=run_scraping, fg_color="#b81d1d", corner_radius=15)
run_button.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

status_label = ctk.CTkLabel(data_extractor_frame, text="Status:", text_color="#ffffff")
status_label.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

# Create a frame for the new tab
claude_frame = ctk.CTkFrame(app, width=300, height=400, fg_color='gray10')
claude_button = ctk.CTkButton(claude_frame, text='Open Claude', command=open_claude, corner_radius=15, fg_color="#b81d1d")
claude_button.grid(row=1, column=2, columnspan=2, padx=10, pady=10)

# Add your new tab content to claude_frame here...

# Show the first tab by default
show_frame(data_extractor_frame, data_extractor_button)

app.geometry("450x550")
# Start GUI loop
app.mainloop()




